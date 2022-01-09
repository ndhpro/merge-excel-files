import os
import openpyxl as xl
from glob import glob
from copy import copy

try:
    OUT_FILE = "Tổng hợp.xlsx"

    if os.path.exists(OUT_FILE):
        os.remove(OUT_FILE)

    file_paths = glob("*.xlsx")

    out_wb = xl.Workbook()
    out_ws = out_wb.active

    is_header = True
    wb = xl.load_workbook(file_paths[0])
    ws = wb.active
    out_r = 1
    for r in range(1, ws.max_row+1):
        if not is_header:
            break
        for c in range(1, ws.max_column+1):
            cell = ws.cell(r, c)
            out_ws.cell(out_r, c).value = cell.value
            if cell.has_style:
                out_ws.cell(out_r, c).fill = copy(cell.fill)
                out_ws.cell(out_r, c).border = copy(cell.border)
                out_ws.cell(out_r, c).protection = copy(
                    cell.protection)
            if cell.value == "(27)":
                is_header = False
                break
        out_r += 1

    for file_path in file_paths:
        wb = xl.load_workbook(file_path, data_only=True)
        ws = wb.active
        max_value = 0
        for r in range(1, ws.max_row+1):
            try:
                value = int(ws.cell(r, 8).value)
            except:
                continue
            if value > max_value:
                for c in range(1, ws.max_column+1):
                    out_ws.cell(out_r, c).value = ws.cell(r, c).value
                max_value = value
        out_r += 1

    out_wb.save(OUT_FILE)
except Exception as e:
    print(e)
    input()
